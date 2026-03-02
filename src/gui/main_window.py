"""
===============================================================================
Project   : Hibi-BuRe
Module    : main_window.py
Created   : 13.02.26
Author    : florian
Purpose   : Main window of the Hibi-BuBe Report Generator application.

@docstyle: google
@language: english
@voice: imperative
===============================================================================
"""

from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QSplitter,
    QMenuBar, QMenu, QStatusBar, QLabel, QFrame, QPushButton, QComboBox
)
from PyQt6.QtCore import Qt, QSize
from PyQt6.QtGui import QAction, QIcon
from pathlib import Path

from src.data.database_manager import DatabaseManager


class MainWindow(QMainWindow):
    """
    Main application window for the Hibi-BuBe Report Generator.

    This class represents the main graphical user interface (GUI) window
    for the application. It handles initialization of the UI components,
    configuration of the window, and setting up the menu bar, central
    widgets, and other elements required for report generation. The
    `MainWindow` serves as a controller for various components, managing
    the application's primary workflow and interactions.

    :ivar db: The database manager that provides access to account and
        category data.
    :type db: DatabaseManager
    """

    def __init__(self, db_manager: DatabaseManager):
        """
        Represents the main application window for the Hibi-BuBe Report Generator, handling the
        user interface, window configuration, theming, and user profile management.

        :param db_manager: Provides access to the database management functionalities.

        :raises TypeError: If db_manager is not an instance of DatabaseManager.
        """
        super().__init__()

        self.db = db_manager

        # Import settings and store as instance variable
        from src.config.settings import settings as _settings
        global settings
        settings = _settings

        # Fenster-Konfiguration
        self.setWindowTitle("Hibi-BuBe Report Generator")
        self.setGeometry(100, 100, 1600, 1000)
        self.setMinimumSize(QSize(1024, 768))

        # UI aufbauen
        self._create_menu_bar()
        self._create_central_widget()
        self._create_status_bar()

        # Theme anwenden
        self._apply_theme()

        # Profile laden
        self._refresh_profile_list()

    def _create_menu_bar(self):
        """
        Creates and configures the main menu bar for the application.

        This method defines and adds multiple menus, such as Datei (File), Bearbeiten
        (Edit), Ansicht (View), and Hilfe (Help), to the application's menu bar. Each
        menu contains various menu items that perform specific actions when triggered.
        Shortcuts and connections to corresponding methods are assigned to these
        actions appropriately.

        :return: None
        """
        menubar = self.menuBar()

        # Datei-Menü
        file_menu = menubar.addMenu("&Datei")

        new_action = QAction("&Neuer Report", self)
        new_action.setShortcut("Ctrl+N")
        new_action.triggered.connect(self._new_report)
        file_menu.addAction(new_action)

        open_action = QAction("&Template laden...", self)
        open_action.setShortcut("Ctrl+O")
        open_action.triggered.connect(self._load_template)
        file_menu.addAction(open_action)

        save_action = QAction("Template &speichern...", self)
        save_action.setShortcut("Ctrl+S")
        save_action.triggered.connect(self._save_template)
        file_menu.addAction(save_action)

        file_menu.addSeparator()

        export_excel = QAction("Als &Excel exportieren...", file_menu)
        export_excel.setShortcut("Ctrl+E")
        export_excel.triggered.connect(self._export_excel)
        file_menu.addAction(export_excel)

        file_menu.addSeparator()

        quit_action = QAction("&Beenden", self)
        quit_action.setShortcut("Ctrl+Q")
        quit_action.triggered.connect(self.close)
        file_menu.addAction(quit_action)

        # Ansicht-Menü
        view_menu = menubar.addMenu("&Ansicht")

        self.dark_mode_action = QAction("&Dark Mode", self, checkable=True)
        self.dark_mode_action.setChecked(settings.app.theme == "dark")
        self.dark_mode_action.triggered.connect(self._toggle_theme)
        view_menu.addAction(self.dark_mode_action)

        view_menu.addSeparator()

        refresh_action = QAction("Aktualisieren", self)
        refresh_action.setShortcut("F5")
        refresh_action.triggered.connect(self._refresh_preview)
        view_menu.addAction(refresh_action)

        # Hilfe-Menü
        help_menu = menubar.addMenu("&Hilfe")

        about_action = QAction("Ü&ber Hibi-BuBe", self)
        about_action.triggered.connect(self._show_about)
        help_menu.addAction(about_action)

    def _create_central_widget(self):
        """
        Creates and sets up the central widget of the main window. The central widget is divided
        into a horizontal layout containing a configurable sidebar and a preview/action area.
        The layout ensures flexible resizing between the sidebar and the main content using
        a splitter.

        :returns: None
        """
        # Central Widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # Hauptlayout (horizontal)
        main_layout = QHBoxLayout(central_widget)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(10)

        # Splitter für anpassbare Breiten
        splitter = QSplitter(Qt.Orientation.Horizontal)

        # === LINKE SIDEBAR: Konfiguration ===
        self.sidebar = self._create_sidebar()
        splitter.addWidget(self.sidebar)

        # === RECHTS: Preview und Actions ===
        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(10)

        # Preview-Bereich
        self.preview_area = self._create_preview_area()
        right_layout.addWidget(self.preview_area, stretch=1)

        # Action-Bereich (unten)
        self.action_area = self._create_action_area()
        right_layout.addWidget(self.action_area)

        splitter.addWidget(right_widget)

        # Splitter-Verhältnis: 25% Sidebar, 75% Preview
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 3)

        main_layout.addWidget(splitter)

    def _create_sidebar(self) -> QWidget:
        """
        Creates and configures a sidebar widget for report configuration in a PyQt application.

        This method sets up a sidebar containing various widgets for report customization. It includes
        input fields, tabs, and modules for account selection, category selection, date range
        selection, field configuration, and grouping configuration. Each section is handled through
        dedicated widgets and is organized in tabs.

        :return: The configured sidebar as a QWidget instance.
        :rtype: QWidget
        """
        from PyQt6.QtWidgets import QTabWidget
        from src.gui.widgets.account_selection_widget import AccountSelectionWidget
        from src.gui.widgets.category_selection_widget import CategorySelectionWidget
        from src.gui.widgets.date_range_widget import DateRangeWidget
        from src.gui.widgets.field_selection_widget import FieldSelectionWidget
        from src.gui.widgets.grouping_config_widget import GroupingConfigWidget

        sidebar = QFrame()
        sidebar.setFrameShape(QFrame.Shape.StyledPanel)
        sidebar.setMinimumWidth(450)
        sidebar.setMaximumWidth(650)

        layout = QVBoxLayout(sidebar)
        layout.setContentsMargins(5, 5, 5, 5)
        layout.setSpacing(5)

        # Überschrift
        title = QLabel("Report-Konfiguration")
        title.setStyleSheet("font-size: 14px; font-weight: bold;")
        layout.addWidget(title)

        # Report-Name Eingabefeld
        name_layout = QHBoxLayout()
        name_label = QLabel("Name:")
        name_label.setStyleSheet("font-weight: bold;")
        name_layout.addWidget(name_label)

        from PyQt6.QtWidgets import QLineEdit
        self.report_name_input = QLineEdit()
        self.report_name_input.setPlaceholderText("Bericht-Name eingeben...")
        self.report_name_input.textChanged.connect(self._on_selection_changed)
        name_layout.addWidget(self.report_name_input)

        layout.addLayout(name_layout)
        layout.addSpacing(5)

        # Tab Widget für die verschiedenen Konfigurationsbereiche
        tabs = QTabWidget()
        tabs.setTabPosition(QTabWidget.TabPosition.North)

        # === Tab 1: Filter (Konten + Kategorien) ===
        filter_tab = QWidget()
        filter_layout = QVBoxLayout(filter_tab)
        filter_layout.setContentsMargins(5, 5, 5, 5)
        filter_layout.setSpacing(10)

        # Konten
        try:
            accounts = self.db.accounts.get_all()
            self.account_widget = AccountSelectionWidget(accounts)
            self.account_widget.selection_changed.connect(self._on_selection_changed)
            filter_layout.addWidget(self.account_widget)
        except Exception as e:
            error_label = QLabel(f"Fehler beim Laden der Konten: {e}")
            error_label.setStyleSheet("color: red;")
            filter_layout.addWidget(error_label)

        filter_layout.addWidget(self._create_separator())

        # Kategorien
        try:
            categories = self.db.categories.get_all()
            self.category_widget = CategorySelectionWidget(categories)
            self.category_widget.selection_changed.connect(self._on_selection_changed)
            filter_layout.addWidget(self.category_widget)
        except Exception as e:
            error_label = QLabel(f"Fehler beim Laden der Kategorien: {e}")
            error_label.setStyleSheet("color: red;")
            filter_layout.addWidget(error_label)

        filter_layout.addStretch()
        tabs.addTab(filter_tab, "📋 Filter")

        # === Tab 2: Zeitraum ===
        date_tab = QWidget()
        date_layout = QVBoxLayout(date_tab)
        date_layout.setContentsMargins(5, 5, 5, 5)

        self.date_range_widget = DateRangeWidget()
        self.date_range_widget.range_changed.connect(self._on_selection_changed)
        date_layout.addWidget(self.date_range_widget)
        date_layout.addStretch()
        tabs.addTab(date_tab, "📅 Zeitraum")

        # === Tab 3: Felder ===
        fields_tab = QWidget()
        fields_layout = QVBoxLayout(fields_tab)
        fields_layout.setContentsMargins(5, 5, 5, 5)

        self.field_widget = FieldSelectionWidget()
        self.field_widget.selection_changed.connect(self._on_selection_changed)
        fields_layout.addWidget(self.field_widget)
        fields_layout.addStretch()
        tabs.addTab(fields_tab, "📊 Felder")

        # === Tab 4: Optionen (Gruppierung + Export) ===
        grouping_tab = QWidget()
        grouping_layout = QVBoxLayout(grouping_tab)
        grouping_layout.setContentsMargins(5, 5, 5, 5)

        self.grouping_widget = GroupingConfigWidget()
        self.grouping_widget.config_changed.connect(self._on_selection_changed)
        grouping_layout.addWidget(self.grouping_widget)

        # Export-Einstellungen
        from PyQt6.QtWidgets import QGroupBox, QLineEdit, QPushButton, QFileDialog
        export_group = QGroupBox("Export-Einstellungen")
        export_layout = QVBoxLayout()

        # Export-Verzeichnis
        export_dir_layout = QHBoxLayout()
        export_dir_label = QLabel("Export-Verzeichnis:")
        export_dir_layout.addWidget(export_dir_label)

        self.export_dir_input = QLineEdit()
        self.export_dir_input.setPlaceholderText(str(settings.app.export_dir))
        # Lade gespeichertes Export-Verzeichnis
        from PyQt6.QtCore import QSettings
        qsettings = QSettings("Hibi-BuBe", "HibiBuBe")
        saved_export_dir = qsettings.value("export_dir", str(settings.app.export_dir))
        self.export_dir_input.setText(saved_export_dir)
        settings.app.export_dir = Path(saved_export_dir)
        export_dir_layout.addWidget(self.export_dir_input)

        browse_btn = QPushButton("📁 Durchsuchen")
        browse_btn.clicked.connect(self._browse_export_dir)
        export_dir_layout.addWidget(browse_btn)

        export_layout.addLayout(export_dir_layout)

        # Speichern-Button
        save_export_btn = QPushButton("💾 Speichern")
        save_export_btn.clicked.connect(self._save_export_settings)
        export_layout.addWidget(save_export_btn)

        export_group.setLayout(export_layout)
        grouping_layout.addWidget(export_group)

        grouping_layout.addStretch()
        tabs.addTab(grouping_tab, "⚙️ Optionen")

        layout.addWidget(tabs)

        return sidebar

    def _create_separator(self) -> QFrame:
        """
        Creates and returns a horizontal line separator as a QFrame object.

        :return: A QFrame object configured as a horizontal line separator with a
                 sunken shadow.
        :rtype: QFrame
        """
        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setFrameShadow(QFrame.Shadow.Sunken)
        return line

    def _create_preview_area(self) -> QWidget:
        """
        Creates and configures the preview area widget.

        This method initializes a QFrame element styled with a panel-like appearance.
        It sets margins for the layout and includes the addition of a `PreviewWidget`
        for displaying preview-related content. The `PreviewWidget` also integrates
        a signal `refresh_requested` which triggers the appropriate handler for
        refresh operations.

        :return: The fully constructed preview area widget.
        :rtype: QWidget
        """
        from src.gui.widgets.preview_widget import PreviewWidget

        preview = QFrame()
        preview.setFrameShape(QFrame.Shape.StyledPanel)

        layout = QVBoxLayout(preview)
        layout.setContentsMargins(5, 5, 5, 5)

        # Preview Widget
        self.preview_widget = PreviewWidget()
        self.preview_widget.refresh_requested.connect(self._on_refresh_preview)
        layout.addWidget(self.preview_widget)

        return preview

    def _create_action_area(self) -> QWidget:
        """
        Creates and configures the action area widget.

        This method initializes and returns a QWidget that contains UI elements such as a dropdown
        to select profiles, buttons to save profiles, and a button for print/export actions.
        The widget is designed with a maximum height of 80 pixels and uses a horizontal layout
        to arrange its child components.

        :returns: A configured QWidget containing the action area UI elements.
        :rtype: QWidget
        """
        action_area = QFrame()
        action_area.setFrameShape(QFrame.Shape.StyledPanel)
        action_area.setMaximumHeight(80)

        layout = QHBoxLayout(action_area)

        # Template-Auswahl Dropdown
        template_label = QLabel("Profil:")
        template_label.setStyleSheet("font-weight: bold;")
        layout.addWidget(template_label)

        self.profile_combo = QComboBox()
        self.profile_combo.setMinimumWidth(250)
        self.profile_combo.setMaximumWidth(400)
        self.profile_combo.addItem("-- Kein Profil geladen --")
        self.profile_combo.currentIndexChanged.connect(self._on_profile_selected)
        layout.addWidget(self.profile_combo)

        # Profil-Buttons
        save_profile_btn = QPushButton("💾 Speichern")
        save_profile_btn.setToolTip("Aktuelles Profil speichern")
        save_profile_btn.clicked.connect(self._save_profile)
        layout.addWidget(save_profile_btn)

        layout.addStretch()

        # Action-Buttons
        print_btn = QPushButton("🖨 Drucken / Export")
        print_btn.setMinimumHeight(40)
        print_btn.clicked.connect(self._show_print_preview)
        layout.addWidget(print_btn)

        return action_area

    def _create_status_bar(self):
        """
        Creates and sets a status bar for the application window with information about the database
        connection status. The displayed message indicates whether the connection to the database
        is successful and provides a masked URL for security purposes. If the connection fails, an
        error message is shown instead.

        :raises None: No exceptions are explicitly raised in this method.
        """
        statusbar = QStatusBar()
        self.setStatusBar(statusbar)

        # Test DB-Verbindung
        if self.db.test_connection():
            status_text = f"✓ Verbunden mit {self.db.url}"
            # Kürze URL für Anzeige (ohne Passwort)
            import re
            status_text_short = re.sub(r'://([^:]+):([^@]+)@', r'://\1:***@', status_text)
            statusbar.showMessage(status_text_short)
        else:
            statusbar.showMessage("✗ Datenbankverbindung fehlgeschlagen")

    def _apply_theme(self):
        """
        Applies the appropriate visual theme to the application interface based on
        the current settings.

        This method checks the application's theme configuration and applies
        either the dark or light theme accordingly.

        :return: None
        :rtype: None
        """
        if settings.app.theme == "dark":
            self._apply_dark_theme()
        else:
            self._apply_light_theme()

    def _apply_dark_theme(self):
        """
        Applies a custom dark theme to the application's stylesheet.

        This method builds and applies a dark theme stylesheet to the application or
        specific UI components, enabling a consistent dark theme appearance across
        various widgets. The styling includes definitions for widgets such as
        `QMainWindow`, `QFrame`, `QLabel`, `QPushButton`, `QCheckBox`, `QComboBox`,
        `QTreeWidget`, and others. It customizes aspects such as background color,
        font color, border styles, hover states, selection behavior, and other
        relevant UI properties.

        :raises ValueError: If the dark theme stylesheet cannot be applied (if
            applicable in an extended version of this code or usage context).

        :return: None
        """
        dark_stylesheet = """
        QMainWindow {
            background-color: #1e1e1e;
            color: #ffffff;
        }
        QFrame {
            background-color: #252525;
            border: 1px solid #3a3a3a;
            border-radius: 5px;
            padding: 10px;
        }
        QLabel {
            color: #ffffff;
            background: transparent;
        }
        QPushButton {
            background-color: #3a3a3a;
            color: #ffffff;
            border: 1px solid #4a4a4a;
            border-radius: 3px;
            padding: 5px 10px;
        }
        QPushButton:hover {
            background-color: #4a4a4a;
            border-color: #5a5a5a;
        }
        QPushButton:pressed {
            background-color: #2a2a2a;
        }
        QCheckBox {
            color: #ffffff;
            background: transparent;
            spacing: 5px;
        }
        QCheckBox::indicator {
            width: 18px;
            height: 18px;
            border: 2px solid #5a5a5a;
            border-radius: 3px;
            background-color: #2a2a2a;
        }
        QCheckBox::indicator:hover {
            border-color: #6a6a6a;
        }
        QCheckBox::indicator:checked {
            background-color: #4caf50;
            border-color: #4caf50;
        }
        QRadioButton {
            color: #ffffff;
            background: transparent;
            spacing: 5px;
        }
        QRadioButton::indicator {
            width: 18px;
            height: 18px;
            border: 2px solid #5a5a5a;
            border-radius: 9px;
            background-color: #2a2a2a;
        }
        QRadioButton::indicator:checked {
            background-color: #4caf50;
            border-color: #4caf50;
        }
        QComboBox {
            background-color: #2a2a2a;
            color: #ffffff;
            border: 1px solid #4a4a4a;
            border-radius: 3px;
            padding: 4px 8px;
            min-height: 20px;
        }
        QComboBox:hover {
            border-color: #5a5a5a;
        }
        QComboBox::drop-down {
            border: none;
            width: 20px;
        }
        QComboBox QAbstractItemView {
            background-color: #2a2a2a;
            color: #ffffff;
            selection-background-color: #1e88e5;
            selection-color: #ffffff;
            border: 1px solid #4a4a4a;
        }
        QDateEdit {
            background-color: #2a2a2a;
            color: #ffffff;
            border: 1px solid #4a4a4a;
            border-radius: 3px;
            padding: 4px;
            min-height: 20px;
        }
        QDateEdit:hover {
            border-color: #5a5a5a;
        }
        QScrollArea {
            border: none;
            background-color: transparent;
        }
        QScrollBar:vertical {
            background-color: #2a2a2a;
            width: 12px;
            border: none;
        }
        QScrollBar::handle:vertical {
            background-color: #4a4a4a;
            border-radius: 6px;
            min-height: 20px;
        }
        QScrollBar::handle:vertical:hover {
            background-color: #5a5a5a;
        }
        QScrollBar:horizontal {
            background-color: #2a2a2a;
            height: 12px;
            border: none;
        }
        QScrollBar::handle:horizontal {
            background-color: #4a4a4a;
            border-radius: 6px;
            min-width: 20px;
        }
        QScrollBar::handle:horizontal:hover {
            background-color: #5a5a5a;
        }
        QScrollBar::add-line, QScrollBar::sub-line {
            background-color: #2a2a2a;
            border: none;
        }
        QScrollBar::add-page, QScrollBar::sub-page {
            background-color: #2a2a2a;
        }
        QScrollBar::left-arrow, QScrollBar::right-arrow,
        QScrollBar::up-arrow, QScrollBar::down-arrow {
            background-color: #4a4a4a;
            border: none;
        }
        QTreeWidget, QTreeView {
            background-color: #1a1a1a;
            color: #ffffff;
            border: 1px solid #3a3a3a;
            alternate-background-color: #202020;
            gridline-color: #3a3a3a;
        }
        QTreeWidget::item, QTreeView::item {
            color: #ffffff;
            padding: 4px;
        }
        QTreeWidget::item:selected, QTreeView::item:selected {
            background-color: #1e88e5;
            color: #ffffff;
        }
        QTreeWidget::item:hover, QTreeView::item:hover {
            background-color: #2a2a2a;
        }
        QTreeWidget::indicator, QTreeView::indicator {
            width: 18px;
            height: 18px;
            border: 2px solid #5a5a5a;
            border-radius: 3px;
            background-color: #2a2a2a;
        }
        QTreeWidget::indicator:hover, QTreeView::indicator:hover {
            border-color: #6a6a6a;
        }
        QTreeWidget::indicator:checked, QTreeView::indicator:checked {
            background-color: #4caf50;
            border-color: #4caf50;
        }
        QHeaderView::section {
            background-color: #2a2a2a;
            color: #ffffff;
            border: 1px solid #3a3a3a;
            padding: 10px 8px;
            font-weight: bold;
            min-height: 40px;
            font-size: 13px;
        }
        QGroupBox {
            color: #ffffff;
            border: 1px solid #4a4a4a;
            border-radius: 5px;
            margin-top: 12px;
            padding-top: 12px;
            font-weight: bold;
        }
        QGroupBox::title {
            subcontrol-origin: margin;
            subcontrol-position: top left;
            padding: 0 5px;
            background-color: transparent;
            color: #ffffff;
        }
        QLineEdit {
            background-color: #2a2a2a;
            color: #ffffff;
            border: 1px solid #4a4a4a;
            border-radius: 3px;
            padding: 4px;
        }
        QLineEdit:hover {
            border-color: #5a5a5a;
        }
        QLineEdit:focus {
            border-color: #1e88e5;
        }
        QMenuBar {
            background-color: #252525;
            color: #ffffff;
        }
        QMenuBar::item {
            background-color: transparent;
            padding: 4px 8px;
        }
        QMenuBar::item:selected {
            background-color: #3a3a3a;
        }
        QMenu {
            background-color: #252525;
            color: #ffffff;
            border: 1px solid #3a3a3a;
        }
        QMenu::item {
            padding: 5px 25px;
        }
        QMenu::item:selected {
            background-color: #1e88e5;
        }
        QStatusBar {
            background-color: #252525;
            color: #ffffff;
        }
        QTabWidget::pane {
            border: 1px solid #3a3a3a;
            background-color: #252525;
            top: -1px;
        }
        QTabBar::tab {
            background-color: #2a2a2a;
            color: #aaaaaa;
            border: 1px solid #3a3a3a;
            padding: 8px 16px;
            margin-right: 2px;
            border-bottom: none;
        }
        QTabBar::tab:selected {
            background-color: #252525;
            color: #ffffff;
            border-bottom-color: #252525;
            font-weight: bold;
        }
        QTabBar::tab:hover {
            background-color: #353535;
            color: #ffffff;
        }
        QTabBar QToolButton {
            background-color: #2a2a2a;
            border: 1px solid #3a3a3a;
            color: #ffffff;
        }
        QTabBar QToolButton:hover {
            background-color: #3a3a3a;
        }
        QTabBar::scroller {
            width: 20px;
        }
        QMessageBox {
            background-color: #2a2a2a;
            color: #ffffff;
        }
        QMessageBox QLabel {
            color: #ffffff;
            background-color: transparent;
        }
        QMessageBox QPushButton {
            background-color: #3a3a3a;
            color: #ffffff;
            border: 1px solid #4a4a4a;
            border-radius: 3px;
            padding: 6px 16px;
            min-width: 60px;
        }
        QMessageBox QPushButton:hover {
            background-color: #4a4a4a;
            border-color: #5a5a5a;
        }
        QMessageBox QPushButton:pressed {
            background-color: #1e88e5;
        }
        """
        self.setStyleSheet(dark_stylesheet)

    def _apply_light_theme(self):
        """
        Applies a predefined light theme to the application by assigning a specific stylesheet.

        This method sets up consistent, visually appealing, and functional styling for
        various Qt widgets. The light theme uses soft background colors, high contrasts
        for readability, and hover/focus effects for improved user experience. This method
        ensures uniform styling across different components including windows, frames, labels,
        checkboxes, radio buttons, drop-downs (comboboxes), date selectors, scrollbars, trees,
        headers, group boxes, line edits, menus, status bars, tabs, message boxes, and more.

        The stylesheet applied supports adaptive interactions like hover states, focus changes,
        selection behaviors, and other UI feedback mechanisms. The theme enhances the application’s
        usability while maintaining a professional and modern look.

        Raises
        ------
        :rtype: None
            No return value is expected from this method.
        """
        light_stylesheet = """
        QMainWindow {
            background-color: #f0f0f0;
            color: #000000;
        }
        QFrame {
            background-color: #ffffff;
            border: 1px solid #cccccc;
            border-radius: 5px;
            padding: 10px;
        }
        QLabel {
            color: #000000;
        }
        QCheckBox {
            color: #000000;
            background: transparent;
            spacing: 5px;
        }
        QCheckBox::indicator {
            width: 18px;
            height: 18px;
            border: 2px solid #999999;
            border-radius: 3px;
            background-color: #ffffff;
        }
        QCheckBox::indicator:hover {
            border-color: #666666;
        }
        QCheckBox::indicator:checked {
            background-color: #4caf50;
            border-color: #4caf50;
        }
        QRadioButton {
            color: #000000;
            background: transparent;
            spacing: 5px;
        }
        QRadioButton::indicator {
            width: 18px;
            height: 18px;
            border: 2px solid #999999;
            border-radius: 9px;
            background-color: #ffffff;
        }
        QRadioButton::indicator:checked {
            background-color: #4caf50;
            border-color: #4caf50;
        }
        QComboBox {
            background-color: #ffffff;
            color: #000000;
            border: 1px solid #cccccc;
            border-radius: 3px;
            padding: 4px 8px;
            min-height: 20px;
        }
        QComboBox:hover {
            border-color: #999999;
        }
        QComboBox::drop-down {
            border: none;
            width: 20px;
        }
        QComboBox QAbstractItemView {
            background-color: #ffffff;
            color: #000000;
            selection-background-color: #1e88e5;
            selection-color: #ffffff;
            border: 1px solid #cccccc;
        }
        QDateEdit {
            background-color: #ffffff;
            color: #000000;
            border: 1px solid #cccccc;
            border-radius: 3px;
            padding: 4px;
            min-height: 20px;
        }
        QDateEdit:hover {
            border-color: #999999;
        }
        QScrollArea {
            border: none;
            background-color: transparent;
        }
        QScrollBar:vertical {
            background-color: #f0f0f0;
            width: 12px;
            border: none;
        }
        QScrollBar::handle:vertical {
            background-color: #cccccc;
            border-radius: 6px;
            min-height: 20px;
        }
        QScrollBar::handle:vertical:hover {
            background-color: #999999;
        }
        QScrollBar:horizontal {
            background-color: #f0f0f0;
            height: 12px;
            border: none;
        }
        QScrollBar::handle:horizontal {
            background-color: #cccccc;
            border-radius: 6px;
            min-width: 20px;
        }
        QScrollBar::handle:horizontal:hover {
            background-color: #999999;
        }
        QScrollBar::add-line, QScrollBar::sub-line {
            background-color: #f0f0f0;
            border: none;
        }
        QScrollBar::add-page, QScrollBar::sub-page {
            background-color: #f0f0f0;
        }
        QScrollBar::left-arrow, QScrollBar::right-arrow,
        QScrollBar::up-arrow, QScrollBar::down-arrow {
            background-color: #cccccc;
            border: none;
        }
        QTreeWidget, QTreeView {
            background-color: #ffffff;
            color: #000000;
            border: 1px solid #cccccc;
            alternate-background-color: #f8f8f8;
            gridline-color: #e0e0e0;
        }
        QTreeWidget::item, QTreeView::item {
            color: #000000;
            padding: 4px;
        }
        QTreeWidget::item:selected, QTreeView::item:selected {
            background-color: #1e88e5;
            color: #ffffff;
        }
        QTreeWidget::item:hover, QTreeView::item:hover {
            background-color: #e8e8e8;
        }
        QTreeWidget::indicator, QTreeView::indicator {
            width: 18px;
            height: 18px;
            border: 2px solid #999999;
            border-radius: 3px;
            background-color: #ffffff;
        }
        QTreeWidget::indicator:hover, QTreeView::indicator:hover {
            border-color: #666666;
        }
        QTreeWidget::indicator:checked, QTreeView::indicator:checked {
            background-color: #4caf50;
            border-color: #4caf50;
        }
        QHeaderView::section {
            background-color: #e8e8e8;
            color: #000000;
            border: 1px solid #cccccc;
            padding: 10px 8px;
            font-weight: bold;
            min-height: 40px;
            font-size: 13px;
        }
        QGroupBox {
            color: #000000;
            border: 1px solid #cccccc;
            border-radius: 5px;
            margin-top: 12px;
            padding-top: 12px;
            font-weight: bold;
        }
        QGroupBox::title {
            subcontrol-origin: margin;
            subcontrol-position: top left;
            padding: 0 5px;
            background-color: transparent;
            color: #000000;
        }
        QLineEdit {
            background-color: #ffffff;
            color: #000000;
            border: 1px solid #cccccc;
            border-radius: 3px;
            padding: 4px;
        }
        QLineEdit:hover {
            border-color: #999999;
        }
        QLineEdit:focus {
            border-color: #1e88e5;
        }
        QMenuBar {
            background-color: #f0f0f0;
            color: #000000;
        }
        QMenuBar::item {
            background-color: transparent;
            padding: 4px 8px;
        }
        QMenuBar::item:selected {
            background-color: #e0e0e0;
        }
        QMenu {
            background-color: #ffffff;
            color: #000000;
            border: 1px solid #cccccc;
        }
        QMenu::item {
            padding: 5px 25px;
        }
        QMenu::item:selected {
            background-color: #1e88e5;
            color: #ffffff;
        }
        QStatusBar {
            background-color: #f0f0f0;
            color: #000000;
        }
        QTabWidget::pane {
            border: 1px solid #cccccc;
            background-color: #ffffff;
            top: -1px;
        }
        QTabBar::tab {
            background-color: #e8e8e8;
            color: #666666;
            border: 1px solid #cccccc;
            padding: 8px 16px;
            margin-right: 2px;
            border-bottom: none;
        }
        QTabBar::tab:selected {
            background-color: #ffffff;
            color: #000000;
            border-bottom-color: #ffffff;
            font-weight: bold;
        }
        QTabBar::tab:hover {
            background-color: #f0f0f0;
            color: #000000;
        }
        QTabBar QToolButton {
            background-color: #e8e8e8;
            border: 1px solid #cccccc;
            color: #000000;
        }
        QTabBar QToolButton:hover {
            background-color: #d8d8d8;
        }
        QTabBar::scroller {
            width: 20px;
        }
        QMessageBox {
            background-color: #ffffff;
            color: #000000;
        }
        QMessageBox QLabel {
            color: #000000;
            background-color: transparent;
        }
        QMessageBox QPushButton {
            background-color: #f0f0f0;
            color: #000000;
            border: 1px solid #cccccc;
            border-radius: 3px;
            padding: 6px 16px;
            min-width: 60px;
        }
        QMessageBox QPushButton:hover {
            background-color: #e0e0e0;
            border-color: #aaaaaa;
        }
        QMessageBox QPushButton:pressed {
            background-color: #1e88e5;
            color: #ffffff;
        }
        """
        self.setStyleSheet(light_stylesheet)

    # === Profile Management ===

    def _refresh_profile_list(self):
        """
        Updates the profile dropdown list by fetching the latest list of profiles from the
        ProfileManager service and repopulating the combo box with the profile data. This
        method temporarily blocks signals during the update to avoid unintended behaviors.

        :raises ImportError: If the 'ProfileManager' module cannot be imported.
        """
        from src.business.services.profile_manager import ProfileManager

        manager = ProfileManager()
        profiles = manager.list_profiles()

        # Blockiere Signale während wir die Liste aktualisieren
        self.profile_combo.blockSignals(True)
        self.profile_combo.clear()

        # Standard-Item
        self.profile_combo.addItem("-- Kein Profil geladen --", None)

        # Füge Profile hinzu
        for profile in profiles:
            self.profile_combo.addItem(profile['name'], profile)

        self.profile_combo.blockSignals(False)

    def _on_profile_selected(self, index: int):
        """
        Handles the event when a profile is selected from a combo box. This function checks
        the selected index and loads the profile configuration if available. The loaded
        configuration is then applied to the UI, and a user notification is displayed.
        If an error occurs during profile loading or application, the user is notified
        via a critical message box.

        :param index: The index of the selected profile in the combo box.
        :type index: int
        :return: None
        """
        if index <= 0:  # "-- Kein Profil geladen --"
            return

        profile_data = self.profile_combo.itemData(index)
        if profile_data:
            try:
                from src.business.services.profile_manager import ProfileManager

                # Lade die vollständige Konfiguration aus der Datei
                manager = ProfileManager()
                config = manager.load_profile(profile_data['filename'])

                # Wende Konfiguration auf UI an
                self._apply_config_to_ui(config)
                self.statusBar().showMessage(f"✓ Profil '{profile_data['name']}' geladen", 3000)
            except Exception as e:
                from PyQt6.QtWidgets import QMessageBox
                QMessageBox.critical(self, "Fehler", f"Fehler beim Laden des Profils:\n{str(e)}")
                import traceback
                traceback.print_exc()

    # === Event Handlers ===

    def _new_report(self):
        """
        Creates a new report and displays a status bar message.

        This method is responsible for initializing a new report and provides feedback
        by displaying a message on the status bar for a limited duration.

        :return: None
        """
        self.statusBar().showMessage("Neuer Report erstellt", 3000)

    def _load_profile(self):
        """
        Handles the loading of user profiles via a dialog interface. If profiles are available,
        displays a dialog allowing the user to select and load a profile. Applies the selected
        profile's configuration to the application's UI and updates the appropriate widgets.
        Supports dark mode styling for the dialog components.

        :param self: The reference to the current instance of the class.

        :raises Exception: If an error occurs during the loading of the selected profile.

        :return: None
        """
        from PyQt6.QtWidgets import QDialog, QVBoxLayout, QListWidget, QDialogButtonBox, QLabel, QMessageBox
        from src.business.services.profile_manager import ProfileManager

        manager = ProfileManager()
        profiles = manager.list_profiles()

        if not profiles:
            QMessageBox.information(self, "Keine Profile", "Es sind noch keine Profile gespeichert.")
            return

        # Dialog erstellen
        dialog = QDialog(self)
        dialog.setWindowTitle("Profil laden")
        dialog.resize(400, 300)

        # Style für Dark Mode
        from src.config.settings import settings
        if settings.app.theme == "dark":
            dialog.setStyleSheet("""
                QDialog {
                    background-color: #1a1a1a;
                    color: #ffffff;
                }
                QLabel {
                    color: #ffffff;
                }
                QListWidget {
                    background-color: #2a2a2a;
                    color: #ffffff;
                    border: 1px solid #3a3a3a;
                }
                QListWidget::item:selected {
                    background-color: #1e88e5;
                    color: #ffffff;
                }
                QListWidget::item:hover {
                    background-color: #3a3a3a;
                }
                QPushButton {
                    background-color: #3a3a3a;
                    color: #ffffff;
                    border: 1px solid #4a4a4a;
                    border-radius: 3px;
                    padding: 6px 16px;
                }
                QPushButton:hover {
                    background-color: #4a4a4a;
                }
                QPushButton:pressed {
                    background-color: #1e88e5;
                }
            """)

        layout = QVBoxLayout(dialog)

        label = QLabel("Wählen Sie ein Profil:")
        layout.addWidget(label)

        # Liste der Profile
        list_widget = QListWidget()
        for profile in profiles:
            list_widget.addItem(profile['name'])
        layout.addWidget(list_widget)

        # Buttons
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)

        # Dialog anzeigen
        if dialog.exec() == QDialog.DialogCode.Accepted and list_widget.currentItem():
            selected_name = list_widget.currentItem().text()

            # Finde Profil
            selected_profile = next((p for p in profiles if p['name'] == selected_name), None)
            if selected_profile:
                try:
                    # Lade Profil
                    config = manager.load_profile(selected_profile['filename'])

                    # Wende Konfiguration auf UI an
                    self._apply_config_to_ui(config)

                    # Setze im Dropdown das geladene Profil
                    for i in range(self.profile_combo.count()):
                        if self.profile_combo.itemText(i) == config.name:
                            self.profile_combo.blockSignals(True)
                            self.profile_combo.setCurrentIndex(i)
                            self.profile_combo.blockSignals(False)
                            break

                    self.statusBar().showMessage(f"✓ Profil '{config.name}' geladen", 3000)
                except Exception as e:
                    QMessageBox.critical(self, "Fehler", f"Fehler beim Laden des Profils:\n{str(e)}")

    def _save_profile(self):
        """
        Saves the current profile configuration. If the profile does not yet have a
        name or is set to the default "Neuer Bericht", a warning will prompt the user
        to provide a name. If a profile with the specified name already exists, the
        user will be prompted to confirm before overwriting it.

        This method handles saving the profile data to persistent storage and updating
        the UI to reflect the changes, such as refreshing the profile list in the
        dropdown menu and setting the newly saved profile as active. In case of an
        error during the save process, an error message is displayed to the user.

        :raises Exception: If there is an issue during the profile saving process, an
            error message is displayed, and the exception is propagated internally.
        """
        from PyQt6.QtWidgets import QMessageBox
        from src.business.services.profile_manager import ProfileManager

        # Hole aktuelle Konfiguration
        config = self.get_report_config()

        # Prüfe ob Name gesetzt ist
        if not config.name or config.name == "Neuer Bericht":
            QMessageBox.warning(self, "Kein Name", "Bitte geben Sie einen Namen für das Profil ein.")
            return

        manager = ProfileManager()

        # Prüfe ob Profil bereits existiert
        if manager.profile_exists(config.name):
            reply = QMessageBox.question(
                self,
                "Profil überschreiben?",
                f"Ein Profil mit dem Namen '{config.name}' existiert bereits.\nMöchten Sie es überschreiben?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.No:
                return

        try:
            # Speichere Profil
            filepath = manager.save_profile(config)
            self.statusBar().showMessage(f"✓ Profil '{config.name}' gespeichert", 3000)

            # Aktualisiere Profil-Liste im Dropdown
            self._refresh_profile_list()

            # Setze das gerade gespeicherte Profil als aktiv im Dropdown
            for i in range(self.profile_combo.count()):
                if self.profile_combo.itemText(i) == config.name:
                    self.profile_combo.blockSignals(True)
                    self.profile_combo.setCurrentIndex(i)
                    self.profile_combo.blockSignals(False)
                    break
        except Exception as e:
            QMessageBox.critical(self, "Fehler", f"Fehler beim Speichern des Profils:\n{str(e)}")

    def _apply_config_to_ui(self, config: 'ReportConfig'):
        """
        Applies the given configuration to the user interface by setting various UI
        components according to the values in the provided `ReportConfig` object.

        This method dynamically checks for the existence of UI components and updates
        their state or values only if they exist. The function ensures that specific
        preferences such as selected accounts, categories, date ranges, fields,
        grouping, and other settings related to printing configurations are correctly
        applied to the corresponding UI widgets.

        :param config: A `ReportConfig` object containing configuration details to be
                       applied to the UI.
        :type config: ReportConfig
        :return: None
        """
        # Name
        if hasattr(self, 'report_name_input'):
            self.report_name_input.setText(config.name)

        # Konten
        if hasattr(self, 'account_widget'):
            self.account_widget.set_selected_account_ids(config.account_ids)

        # Kategorien
        if hasattr(self, 'category_widget'):
            self.category_widget.set_selected_category_ids(config.category_ids)

        # Zeitraum
        if hasattr(self, 'date_range_widget') and config.date_range:
            self.date_range_widget.set_date_range(config.date_range)

        # Felder
        if hasattr(self, 'field_widget'):
            self.field_widget.set_selected_fields(config.fields)
            self.field_widget.set_show_count(config.show_count)

        # Gruppierung
        if hasattr(self, 'grouping_widget'):
            self.grouping_widget.set_config({
                'grouping': config.grouping,
                'sort_field': config.sort_field,
                'sort_order': config.sort_order,
                'show_subtotals': config.show_subtotals,
                'show_grand_total': config.show_grand_total
            })

        # Spalteneinstellungen für Print Preview
        if hasattr(self, 'preview_widget') and config.column_order and config.column_widths:
            self.preview_widget._saved_column_order = config.column_order.copy()
            self.preview_widget._saved_column_widths = config.column_widths.copy()

    def _load_template(self):
        """
        Loads the template content for the application. This method ensures that the
        necessary profile settings are loaded before retrieving the template.

        :return: None
        :rtype: None
        """
        self._load_profile()

    def _save_template(self):
        """
        Saves the template by invoking the private method `_save_profile`.
        This method is used internally and should not be accessed or modified
        externally.

        :return: None
        """
        self._save_profile()

    def _get_file_dialog_stylesheet(self) -> str:
        """
        Returns a comprehensive stylesheet for QFileDialog in dark mode.
        
        :return: Dark mode stylesheet for file dialogs
        :rtype: str
        """
        return """
        QFileDialog {
            background-color: #1e1e1e;
            color: #ffffff;
        }
        QFileDialog QWidget {
            background-color: #1e1e1e;
            color: #ffffff;
        }
        QFileDialog QListView, QFileDialog QTreeView, QFileDialog QTableView {
            background-color: #252525;
            color: #ffffff;
            border: 1px solid #3a3a3a;
            alternate-background-color: #2a2a2a;
        }
        QFileDialog QListView::item:selected, QFileDialog QTreeView::item:selected {
            background-color: #1e88e5;
            color: #ffffff;
        }
        QFileDialog QListView::item:hover, QFileDialog QTreeView::item:hover {
            background-color: #2a2a2a;
        }
        QFileDialog QPushButton {
            background-color: #3a3a3a;
            color: #ffffff;
            border: 1px solid #4a4a4a;
            border-radius: 3px;
            padding: 5px 15px;
            min-width: 60px;
        }
        QFileDialog QPushButton:hover {
            background-color: #4a4a4a;
            border-color: #5a5a5a;
        }
        QFileDialog QPushButton:pressed {
            background-color: #2a2a2a;
        }
        QFileDialog QLineEdit, QFileDialog QTextEdit {
            background-color: #2a2a2a;
            color: #ffffff;
            border: 1px solid #4a4a4a;
            border-radius: 3px;
            padding: 4px;
        }
        QFileDialog QLineEdit:focus {
            border-color: #1e88e5;
        }
        QFileDialog QComboBox {
            background-color: #2a2a2a;
            color: #ffffff;
            border: 1px solid #4a4a4a;
            border-radius: 3px;
            padding: 4px 8px;
        }
        QFileDialog QComboBox:hover {
            border-color: #5a5a5a;
        }
        QFileDialog QComboBox::drop-down {
            border: none;
            width: 20px;
        }
        QFileDialog QComboBox QAbstractItemView {
            background-color: #2a2a2a;
            color: #ffffff;
            selection-background-color: #1e88e5;
            selection-color: #ffffff;
            border: 1px solid #4a4a4a;
        }
        QFileDialog QLabel {
            color: #ffffff;
            background: transparent;
        }
        QFileDialog QHeaderView::section {
            background-color: #2a2a2a;
            color: #ffffff;
            border: 1px solid #3a3a3a;
            padding: 4px;
        }
        QFileDialog QScrollBar:vertical {
            background-color: #2a2a2a;
            width: 12px;
        }
        QFileDialog QScrollBar::handle:vertical {
            background-color: #4a4a4a;
            border-radius: 6px;
        }
        QFileDialog QScrollBar::handle:vertical:hover {
            background-color: #5a5a5a;
        }
        QFileDialog QScrollBar:horizontal {
            background-color: #2a2a2a;
            height: 12px;
        }
        QFileDialog QScrollBar::handle:horizontal {
            background-color: #4a4a4a;
            border-radius: 6px;
        }
        QFileDialog QSplitter::handle {
            background-color: #3a3a3a;
        }
        """

    def _export_excel(self):
        """
        Exports transaction data to an Excel file. This method handles user interaction
        for selecting a file location and name and ensures the transaction data is
        available before proceeding. It also handles potential exceptions during the
        export process.

        :raises AttributeError: If the preview widget does not have the 'transactions'
            attribute or the attribute is empty.
        :raises Exception: If any exception occurs during the export process.

        :return: None
        """
        from PyQt6.QtWidgets import QFileDialog, QMessageBox

        # Prüfe ob Daten vorhanden
        if not hasattr(self.preview_widget, 'transactions') or not self.preview_widget.transactions:
            QMessageBox.warning(
                self,
                "Keine Daten",
                "Bitte laden Sie zuerst Daten mit dem '⟳ Aktualisieren' Button."
            )
            return

        # Dateiname vorschlagen
        config = self.preview_widget.config
        safe_name = config.name.replace('/', '_').replace('\\', '_').replace(':', '_')
        suggested_name = f"{safe_name}.xlsx"

        # Vollständiger Pfad mit Export-Verzeichnis
        from pathlib import Path
        default_path = settings.app.export_dir / suggested_name

        # Datei-Dialog mit Dark Mode Support
        dialog = QFileDialog(self)
        dialog.setWindowTitle("Excel-Export")
        dialog.setAcceptMode(QFileDialog.AcceptMode.AcceptSave)
        dialog.setNameFilter("Excel-Dateien (*.xlsx);;Alle Dateien (*)")
        dialog.setDefaultSuffix("xlsx")
        dialog.selectFile(str(default_path))
        
        # Wende Theme auf Dialog an
        if settings.app.theme == "dark":
            dialog.setStyleSheet(self._get_file_dialog_stylesheet())
        
        if dialog.exec() != QFileDialog.DialogCode.Accepted:
            return
        
        selected_files = dialog.selectedFiles()
        if not selected_files:
            return
            
        filename = selected_files[0]

        # Stelle sicher dass .xlsx Extension vorhanden
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'

        try:
            # Exportiere
            self._export_to_excel(filename)
            QMessageBox.information(self, "Excel exportiert", f"Excel wurde exportiert:\n{filename}")
            self.statusBar().showMessage(f"✓ Excel exportiert: {filename}", 3000)
        except Exception as e:
            QMessageBox.critical(self, "Fehler", f"Fehler beim Excel-Export:\n{str(e)}")
            import traceback
            traceback.print_exc()

    def _export_to_excel(self, filename: str):
        """
        Exports data to an Excel file.

        This method generates and saves a structured Excel report using the 'openpyxl' library.
        It fetches data from the preview widget, including configurations, transactions, headers,
        and relevant tree structures, and organizes it into a hierarchical or flat format based
        on the current grouping configuration.

        :param filename: The path where the Excel file will be saved
        :type filename: str
        :return: None
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.critical(
                self,
                "Bibliothek fehlt",
                "Die Bibliothek 'openpyxl' ist nicht installiert.\n\n"
                "Installieren Sie sie mit:\npip install openpyxl"
            )
            return

        # Erstelle Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"

        config = self.preview_widget.config
        transactions = self.preview_widget.transactions
        tree = self.preview_widget.tree

        # Header-Info
        ws['A1'] = config.name
        ws['A1'].font = Font(size=16, bold=True)

        date_range_str = ""
        if config.date_range:
            date_range_str = f"{config.date_range.start_date.strftime('%d.%m.%Y')} - {config.date_range.end_date.strftime('%d.%m.%Y')}"
        ws['A2'] = f"Zeitraum: {date_range_str}"
        ws['A3'] = f"Buchungen: {len(transactions)}"

        # Hole Header aus Preview (mit Spaltenreihenfolge)
        headers = self.preview_widget.get_headers()

        # Schreibe Header (Zeile 5)
        header_row = 5
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Hole Daten (hierarchisch oder flach)
        from src.gui.dialogs.print_preview_dialog import PrintPreviewDialog
        dialog = PrintPreviewDialog(
            config=config,
            transactions=transactions,
            tree=tree,
            headers=headers,
            column_order=self.preview_widget.get_column_order(),
            column_widths=self.preview_widget.get_column_widths(),
            parent=self
        )

        # Flache Daten aus _flatten_tree
        if config.grouping and tree:
            rows_data = dialog._flatten_tree(headers)
        else:
            rows_data = []
            for trans in transactions:
                row_data = dialog._create_transaction_row_data(trans, headers)
                rows_data.append(("transaction", row_data))

        # Schreibe Daten
        current_row = header_row + 1
        for row_type, row_data in rows_data:
            # Überspringe Header und Separator
            if row_type in ["category_header", "subcategory_header", "separator"]:
                # Füge Kategorie/Unterkategorie als separate Zeile ein
                if row_type in ["category_header", "subcategory_header"]:
                    cell = ws.cell(row=current_row, column=1, value=row_data[0])
                    cell.font = Font(bold=True, size=12 if row_type == "category_header" else 11)
                    cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                    current_row += 1
                elif row_type == "separator":
                    current_row += 1  # Leerzeile
                continue

            # Schreibe Daten-Zeile
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)

                # Styling basierend auf Typ
                if row_type == "subcategory_total":
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                elif row_type == "category_total":
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="A5D6A7", end_color="A5D6A7", fill_type="solid")
                elif row_type == "total":
                    cell.font = Font(bold=True, size=11)
                    cell.fill = PatternFill(start_color="81C784", end_color="81C784", fill_type="solid")

                # Rechts-Ausrichtung für Betrag und Anzahl
                if col_idx <= len(headers) and ("Betrag" in headers[col_idx-1] or headers[col_idx-1] == "Anzahl"):
                    cell.alignment = Alignment(horizontal='right')

            current_row += 1

        # Spaltenbreiten anpassen
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15

        # Speichern
        wb.save(filename)

    def _export_pdf(self):
        """
        Triggers the export of data to a PDF file. This function performs
        necessary operations to initiate a PDF export and displays a status
        message in the application's status bar for a short duration.

        :return: None
        :rtype: NoneType
        """
        self.statusBar().showMessage("PDF-Export (TODO: Stufe 3.2)", 3000)

    def _toggle_theme(self):
        """
        Toggles the application theme between light and dark mode based on the user's
        selection. Updates the application settings accordingly and applies the selected
        theme. A notification is temporarily displayed in the status bar indicating
        the activated mode.
        """
        if self.dark_mode_action.isChecked():
            settings.app.theme = "dark"
            self._apply_dark_theme()
            self.statusBar().showMessage("Dark Mode aktiviert", 2000)
        else:
            settings.app.theme = "light"
            self._apply_light_theme()
            self.statusBar().showMessage("Light Mode aktiviert", 2000)

    def get_report_config(self):
        """
        Generates and returns a report configuration object containing various settings for a report, such as selected
        accounts, categories, fields, date range, column layout, grouping, and sorting configurations.

        :returns: A `ReportConfig` object with the settings extracted from the respective widgets and input fields.
        :rtype: ReportConfig
        """
        from src.data.models.report_config import ReportConfig

        # Hole Werte von Widgets
        # Fallback zu None (keine Filterung) statt [] (leere Liste = nichts anzeigen)
        account_ids = self.account_widget.get_selected_account_ids() if hasattr(self, 'account_widget') else None
        category_ids = self.category_widget.get_selected_category_ids() if hasattr(self, 'category_widget') else None
        date_range = self.date_range_widget.get_date_range() if hasattr(self, 'date_range_widget') else None
        fields = self.field_widget.get_selected_fields() if hasattr(self, 'field_widget') else ["Datum", "Betrag", "Kategorie"]

        # Gruppierungs-Config
        grouping_config = self.grouping_widget.get_config() if hasattr(self, 'grouping_widget') else {}

        # show_count kommt jetzt vom field_widget
        show_count = self.field_widget.get_show_count() if hasattr(self, 'field_widget') else False

        # Report-Name
        report_name = self.report_name_input.text() if hasattr(self, 'report_name_input') and self.report_name_input.text() else "Neuer Bericht"

        # Spalten-Layout aus Preview-Widget (falls vorhanden)
        column_order = []
        column_widths = []
        if hasattr(self, 'preview_widget') and hasattr(self.preview_widget, 'config') and self.preview_widget.config:
            try:
                column_order = self.preview_widget.get_column_order()
                column_widths = self.preview_widget.get_column_widths()
            except:
                pass

        # Erstelle Config
        config = ReportConfig(
            name=report_name,
            account_ids=account_ids,
            category_ids=category_ids,
            date_range=date_range,
            fields=fields,
            grouping=grouping_config.get("grouping", []),
            sort_field=grouping_config.get("sort_field", "Datum"),
            sort_order=grouping_config.get("sort_order", "asc"),
            show_subtotals=grouping_config.get("show_subtotals", True),
            show_grand_total=grouping_config.get("show_grand_total", True),
            show_count=show_count,
            show_debit_credit=grouping_config.get("show_debit_credit", False),
            column_order=column_order,
            column_widths=column_widths
        )

        return config

    def _on_selection_changed(self, *args):
        """
        Handles selection change events, validates the report configuration,
        updates the status bar with the current report details, and refreshes
        the preview if the configuration is valid.

        :param args: Additional arguments passed to the event handler. These
            arguments are not explicitly used within this method.
        """
        # Erstelle Config
        config = self.get_report_config()

        # Validierung
        is_valid, error = config.validate()

        # Status-Update (mit Fallback für None-Werte)
        msg = f"Konten: {len(config.account_ids) if config.account_ids else 0}, "
        msg += f"Kategorien: {len(config.category_ids) if config.category_ids else 0}, "
        msg += f"Felder: {len(config.fields) if config.fields else 0}, "
        msg += f"Gruppierung: {len(config.grouping) if config.grouping else 0}"

        if not is_valid:
            msg += f" | ⚠ {error}"

        self.statusBar().showMessage(msg, 3000)

        # Preview automatisch aktualisieren
        if is_valid:
            self._on_refresh_preview()

    def _on_refresh_preview(self):
        """
        Triggered to refresh the preview by reloading and processing transactions based on the
        current report configuration. This function validates the provided report configuration,
        shows a loading cursor, fetches transactions and categories if required, constructs a
        hierarchical structure, and updates the preview widget with the resulting data. Handles
        errors during processing and restores the UI state.

        :return: None
        """
        config = self.get_report_config()

        # Validierung
        is_valid, error = config.validate()
        if not is_valid:
            self.statusBar().showMessage(f"⚠ Konfiguration ungültig: {error}", 5000)
            return

        # Zeige Wartecursor und Status
        from PyQt6.QtWidgets import QApplication
        from PyQt6.QtCore import Qt

        QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
        self.statusBar().showMessage("⏳ Lade Daten...")
        QApplication.processEvents()

        # Transaktionen laden
        try:
            from src.business.services.hierarchy_builder import HierarchyBuilder

            # WICHTIG:
            # None bedeutet "keine Filterung" = alle anzeigen
            # Leere Liste bedeutet "nichts ausgewählt" = keine Daten anzeigen
            # Verwende direkt die Werte aus config (None oder Liste)
            account_filter = config.account_ids
            category_filter = config.category_ids

            transactions = self.db.transactions.get_by_date_range(
                date_range=config.date_range,
                account_ids=account_filter,
                category_ids=category_filter
            )

            # Kategorien laden (für Tree-Building)
            if config.grouping:
                categories = self.db.categories.get_all()
                builder = HierarchyBuilder()
                tree = builder.build_tree(categories)
                tree = builder.assign_transactions(tree, transactions, categories)
                tree = builder.calculate_aggregates(tree, transactions)
            else:
                tree = None

            # Preview aktualisieren
            self.preview_widget.set_data(config, transactions, tree)

            QApplication.restoreOverrideCursor()
            self.statusBar().showMessage(f"✓ {len(transactions)} Buchungen geladen", 3000)

        except Exception as e:
            QApplication.restoreOverrideCursor()
            self.statusBar().showMessage(f"✗ Fehler beim Laden: {str(e)}", 5000)
            import traceback
            traceback.print_exc()

    def _show_print_preview(self):
        """
        Responsible for displaying the print preview dialog and handling the column
        settings if the dialog is accepted. It verifies the presence of transaction
        data before proceeding and opens a custom print preview dialog for the user
        to confirm or modify settings.

        Raises a warning message box if no transaction data is available.

        :param self: Reference to the containing class instance.

        :raises QMessageBox.warning: Raised when there is no transaction data
            available to show in the print preview dialog.

        :return: None
        """
        # Hole aktuelle Daten vom Preview-Widget
        if not hasattr(self.preview_widget, 'transactions') or not self.preview_widget.transactions:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.warning(
                self,
                "Keine Daten",
                "Bitte laden Sie zuerst Daten mit dem '⟳ Aktualisieren' Button."
            )
            return

        # Öffne PrintPreviewDialog mit gefilterten Transaktionen
        from src.gui.dialogs.print_preview_dialog import PrintPreviewDialog

        # Hole nur die sichtbaren/gefilterten Transaktionen
        filtered_transactions = self.preview_widget.get_filtered_transactions()

        # Debug-Info
        total_count = len(self.preview_widget.transactions)
        filtered_count = len(filtered_transactions)
        search_text = self.preview_widget.search_field.text() if hasattr(self.preview_widget, 'search_field') else ""
        print(f"DEBUG: Total transactions: {total_count}, Filtered: {filtered_count}, Search: '{search_text}'")

        # Wenn gefiltert wurde, baue einen neuen Tree nur mit gefilterten Transaktionen
        # Ansonsten verwende den bestehenden Tree
        tree_to_use = None
        if filtered_count < total_count and self.preview_widget.tree:
            # Filter ist aktiv - baue neuen Tree nur mit gefilterten Transaktionen
            from src.business.services.hierarchy_builder import HierarchyBuilder
            builder = HierarchyBuilder()

            # Hole alle Kategorien
            categories = self._get_all_categories()

            # Baue Tree neu mit gefilterten Transaktionen
            tree_to_use = builder.build_tree(categories)
            tree_to_use = builder.assign_transactions(tree_to_use, filtered_transactions, categories)
            tree_to_use = builder.calculate_aggregates(tree_to_use, filtered_transactions)
        else:
            # Kein Filter aktiv - verwende bestehenden Tree
            tree_to_use = self.preview_widget.tree

        dialog = PrintPreviewDialog(
            config=self.preview_widget.config,
            transactions=filtered_transactions,
            tree=tree_to_use,
            headers=self.preview_widget.get_headers(),
            column_order=self.preview_widget.get_column_order(),
            column_widths=self.preview_widget.get_column_widths(),
            parent=self
        )

        # Dialog öffnen
        result = dialog.exec()

        # Wenn Dialog geschlossen wurde (mit OK), Spalteneinstellungen übernehmen
        if result == dialog.DialogCode.Accepted:
            column_order, column_widths = dialog.get_column_settings()
            if column_order and column_widths:
                # Speichere Einstellungen in PreviewWidget für nächsten Aufruf
                self.preview_widget._saved_column_order = column_order
                self.preview_widget._saved_column_widths = column_widths

    def _get_all_categories(self):
        """
        Hole alle Kategorien aus der Datenbank.

        :return: Liste aller Kategorien
        """
        try:
            return self.db.categories.get_all()
        except Exception as e:
            print(f"Fehler beim Laden der Kategorien: {e}")
            return []

    def _refresh_preview(self):
        """
        Triggers the refresh process for the preview component by invoking a
        helper method to handle the actual refresh operation.
        """
        self._on_refresh_preview()

    def _browse_export_dir(self):
        """
        Opens a directory selection dialog for choosing the export directory.
        Updates the export_dir_input field with the selected path.
        """
        from PyQt6.QtWidgets import QFileDialog
        current_dir = self.export_dir_input.text() or str(settings.app.export_dir)

        # Erstelle Dialog mit Dark Mode Support
        dialog = QFileDialog(self)
        dialog.setWindowTitle("Export-Verzeichnis wählen")
        dialog.setFileMode(QFileDialog.FileMode.Directory)
        dialog.setOption(QFileDialog.Option.ShowDirsOnly, True)
        dialog.setDirectory(current_dir)

        # Wende Theme auf Dialog an
        if settings.app.theme == "dark":
            dialog.setStyleSheet(self._get_file_dialog_stylesheet())

        if dialog.exec() == QFileDialog.DialogCode.Accepted:
            selected_dirs = dialog.selectedFiles()
            if selected_dirs:
                self.export_dir_input.setText(selected_dirs[0])

    def _save_export_settings(self):
        """
        Saves the export directory setting to QSettings and updates
        the global settings object.
        """
        from PyQt6.QtCore import QSettings
        from pathlib import Path

        export_dir = self.export_dir_input.text()
        if not export_dir:
            export_dir = str(settings.app.export_dir)

        # Speichere in QSettings
        qsettings = QSettings("Hibi-BuBe", "HibiBuBe")
        qsettings.setValue("export_dir", export_dir)

        # Update globale Settings
        settings.app.export_dir = Path(export_dir)

        # Stelle sicher, dass das Verzeichnis existiert
        settings.app.export_dir.mkdir(parents=True, exist_ok=True)

        self.statusBar().showMessage(f"✓ Export-Verzeichnis gespeichert: {export_dir}", 3000)

    def _show_about(self):
        """
        Displays an about dialog for the Hibi-BuBe Report Generator application.

        This method utilizes `QMessageBox.about` to present a dialog detailing
        information about the application, including its name, version, framework,
        database, and developer information.

        :return: None
        """
        from PyQt6.QtWidgets import QMessageBox
        QMessageBox.about(
            self,
            "Über Hibi-BuBe",
            "<h2>Hibi-BuBe Report Generator</h2>"
            "<p>Version 0.8 (Beta Version RC 3)</p>"
            "<p>Moderne Report-Applikation für jameica-hibiscus</p>"
            "<p><b>Entwicklung:</b> Florian Mösch</p>"
            "<p><b>Framework:</b> PyQt6</p>"
            "<p><b>Datenbank:</b> MySQL/MariaDB</p>"
        )

    def closeEvent(self, event):
        """
        Closes the database connection and accepts the close event.

        :param event: The close event that signals the window is about to close
        :type event: QCloseEvent
        :return: None
        """
        # DB-Verbindungen schließen
        self.db.close()
        event.accept()


if __name__ == "__main__":
    import sys
    from PyQt6.QtWidgets import QApplication

    print("=== MainWindow Test ===\n")

    # Erstelle Application
    app = QApplication(sys.argv)

    # Erstelle DatabaseManager
    from src.data.database_manager import DatabaseManager
    db = DatabaseManager()

    # Erstelle und zeige MainWindow
    window = MainWindow(db)
    window.show()

    print("✓ MainWindow geöffnet")
    print("✓ Teste Menüs und Theme-Wechsel")

    sys.exit(app.exec())
